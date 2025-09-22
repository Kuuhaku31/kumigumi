# excel_reader.py

from openpyxl import Workbook, load_workbook


class ExcelReader:
    def __init__(己, excel_path: str):

        己.work_book: Workbook = load_workbook(excel_path, data_only=True)

        己.数据库地址: str = ""
        己.更新数据参数: list[tuple[str, str]] = []  # [(数据库表名, 工作表名), ...]
        己.获取数据参数: list[str] = []  # [工作表名, ...]
        己.下载种子参数: list[tuple[str, str]] = []  # [(工作表名, 种子状态), ...]

        己.数据库anime表名: str = ""
        己.数据库episode表名: str = ""
        己.数据库torrent表名: str = ""

        己.种子下载地址: str = ""  # 种子下载地址

        # 解析 main 工作表
        main_sheet = 己.work_book["main"]

        行指针: int = 1
        列指针: int = 1
        while True:
            指令 = main_sheet.cell(行指针, 列指针).value

            if 指令 == "_end":
                break
            elif 指令 == "_to":  # 跳到指定行
                行指针_to = int(main_sheet.cell(行指针, 列指针 + 1).value)
                列指针_to = int(main_sheet.cell(行指针, 列指针 + 2).value)
                行指针 = 行指针_to
                列指针 = 列指针_to
                continue
            elif 指令 is None:
                pass

            elif 指令 == "_database_path":
                己.数据库地址 = main_sheet.cell(行指针, 列指针 + 1).value

            elif 指令 == "_store":
                数据库表名 = main_sheet.cell(行指针, 列指针 + 1).value
                工作表名 = main_sheet.cell(行指针, 列指针 + 2).value
                己.更新数据参数.append((数据库表名, 工作表名))

            elif 指令 == "_fetch":
                工作表名 = main_sheet.cell(行指针, 列指针 + 1).value
                己.获取数据参数.append(工作表名)

            elif 指令 == "_dt":
                工作表名 = main_sheet.cell(行指针, 列指针 + 1).value
                种子下载状态 = main_sheet.cell(行指针, 列指针 + 2).value
                己.下载种子参数.append((工作表名, 种子下载状态))

            elif 指令 == "_db_anime":
                己.数据库anime表名 = main_sheet.cell(行指针, 列指针 + 1).value
            elif 指令 == "_db_episode":
                己.数据库episode表名 = main_sheet.cell(行指针, 列指针 + 1).value
            elif 指令 == "_db_torrent":
                己.数据库torrent表名 = main_sheet.cell(行指针, 列指针 + 1).value

            elif 指令 == "_dt_path":
                己.种子下载地址 = main_sheet.cell(行指针, 列指针 + 1).value

            行指针 += 1

    def 获取工作表数据(己, 工作表名: str) -> list[list[str]]:
        """
        获取指定工作表的数据
        返回一个二维字符串列表
        第一个元素是表头
        表头的第一个字符串是主键
        其他字符串是字段名
        """

        己.ExcelReaderPrint(f"🔄 读取 {工作表名} ...", "")

        工作表 = 己.work_book[工作表名]

        起始行: int = 0
        结束行: int = 0
        主键: str = ""
        字段字典: dict[str, int] = {}  # 字段名 : 列号

        指令: str = ""
        行指针: int = 1
        while True:
            键: str = 工作表.cell(row=行指针, column=1).value
            值: str = 工作表.cell(row=行指针, column=2).value

            if 键 is None:
                pass
            elif 指令 == "_to":  # 跳到指定行
                行指针 = int(值)
                continue
            elif 键 == "_end":
                break
            elif 键 == "_start_row":
                起始行 = int(值)
            elif 键 == "_end_row":
                结束行 = int(值)
            elif 键 == "_primary_key":
                主键 = 值
            else:
                字段字典[键] = int(值)

            行指针 += 1

        if not 主键 or 主键 not in 字段字典:
            raise ValueError(f"❌ 工作表 {工作表名} 中未定义主键或主键 {主键} 不存在")

        # 读取数据区域
        data: list[list[str]] = []
        data_header: list[str] = [主键] + [
            字段名 for 字段名 in 字段字典.keys() if 字段名 != 主键
        ]  # 设置表头，主键在第一个
        data.append(data_header)
        for 行号 in range(起始行, 结束行):
            # 读取每一行的数据
            # 直接从数据区域读取，跳过表头
            row_data: list[str] = []
            for 字段名 in data_header:
                单元格值 = 工作表.cell(row=行号, column=字段字典[字段名]).value
                row_data.append(单元格值 if 单元格值 is not None else "")
            data.append(row_data)

        print(" 读取结束")
        return data

    def 读取sheet获取bgm_url_rss_映射(己, 工作表名: str) -> dict[str, str]:
        """
        读取指定工作表的 bgm_url 和 rss_url 映射
        """

        data = 己.获取工作表数据(工作表名)

        # 获取目标键的列数
        column_bgm_url = -1
        column_rss_url = -1
        for i, header in enumerate(data[0]):
            if header == "anime_bangumi_url":
                column_bgm_url = i
            elif header == "anime_rss_url":
                column_rss_url = i
        if column_bgm_url == -1 or column_rss_url == -1:
            raise ValueError(f"❌ 工作表 {工作表名} 中未找到 'anime_bangumi_url' 或 'anime_rss_url' 列")

        # 构建 bgm_url 和 rss_url 的映射
        bgm_url_rss_映射: dict[str, str] = {}
        for row in data[1:]:
            bgm_url = row[column_bgm_url] if len(row) > column_bgm_url and row[column_bgm_url] else ""
            rss_url = row[column_rss_url] if len(row) > column_rss_url and row[column_rss_url] else ""
            bgm_url_rss_映射[bgm_url] = rss_url

        return bgm_url_rss_映射

    def 获取下载种子url列表(己) -> list[str]:

        torrent_download_url_list: list[str] = []

        for 工作表名, 种子状态 in 己.下载种子参数:

            # 下载种子链接
            if 工作表名 != "":
                己.ExcelReaderPrint(f"🔄 获取 {工作表名} 需要下载的种子链接...")

                # 获取种子下载链接工作表
                sheet_download_torrent = 己.work_book[工作表名]

                起始行: int = 0
                结束行: int = 0
                种子下载链接_column: int = 0
                种子下载情况_column: int = 0

                # 读取种子下载链接
                行指针 = 1
                while True:
                    指令 = sheet_download_torrent.cell(行指针, 1).value

                    if 指令 == "_end":
                        break
                    elif 指令 == "_to":
                        行指针 = int(sheet_download_torrent.cell(行指针, 2).value)
                        continue
                    elif 指令 is None:
                        pass

                    elif 指令 == "_start_row":
                        起始行 = int(sheet_download_torrent.cell(行指针, 2).value)
                    elif 指令 == "_end_row":
                        结束行 = int(sheet_download_torrent.cell(行指针, 2).value)

                    elif 指令 == "torrent_download_url":
                        种子下载链接_column = sheet_download_torrent.cell(行指针, 2).value
                    elif 指令 == "torrent_download_status":
                        种子下载情况_column = sheet_download_torrent.cell(行指针, 2).value

                    行指针 += 1

                # 读取种子下载链接
                for 行号 in range(起始行, 结束行):
                    torrent_download_url = sheet_download_torrent.cell(行号, 种子下载链接_column).value
                    torrent_download_status = sheet_download_torrent.cell(行号, 种子下载情况_column).value
                    if torrent_download_status == 种子状态:
                        torrent_download_url_list.append(torrent_download_url)

        return torrent_download_url_list

    def ExcelReaderPrint(己, message: str, end: str = "\n", *, strip_unencodable: bool = True) -> None:
        import sys

        prefix = "\033[92m[ExcelReader]\033[0m: "
        try:
            print(f"{prefix}{message}", end=end)
        except UnicodeEncodeError:
            encoding = (getattr(sys.stdout, "encoding", None) or "utf-8").lower()
            if strip_unencodable:
                try:
                    safe = message.encode(encoding, errors="ignore").decode(encoding, errors="ignore")
                except Exception:
                    safe = message.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore")
            else:
                try:
                    safe = message.encode(encoding, errors="replace").decode(encoding, errors="replace")
                except Exception:
                    safe = message
            print(f"{prefix}{safe}", end=end)
