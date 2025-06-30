# main.py

import os
import shutil
import tempfile
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Tuple

import bangumi
import headers
import mikananime.mikananime as mikananime
import pyodbc
import requests
import utils
from openpyxl import load_workbook
from tqdm import tqdm
from utils import kumigumiPrint


def 批量下载种子(种子下载链接列表: list[str]):
    """
    批量下载种子
    :param 种子下载链接列表: 包含多个种子下载链接的列表
    :return: None
    """

    download_path = utils.获取用户默认下载路径() + "/dt/"
    os.makedirs(download_path, exist_ok=True)

    # 单个种子下载函数
    def download_torrent(种子下载链接: str):
        try:
            file_name = os.path.basename(种子下载链接)
            file_path = os.path.join(download_path, file_name)
            resp = requests.get(种子下载链接, timeout=30)
            resp.raise_for_status()
            with open(file_path, "wb") as f:
                f.write(resp.content)
        except Exception as e:
            raise RuntimeError(f"下载失败: {种子下载链接}，原因: {e}")

    fail_url_list = []

    # 使用多线程批量下载种子
    with ThreadPoolExecutor() as executor:
        futures = {executor.submit(download_torrent, url): url for url in 种子下载链接列表}

        for future in tqdm(as_completed(futures), total=len(futures), desc="下载种子文件进度"):
            url = futures[future]
            try:
                future.result()  # 等待下载完成
            except Exception as e:
                print(f"❌ 下载种子 {url} 时发生错误: {e}")
                fail_url_list.append(url)

    # 将下载失败的链接保存到文件
    if fail_url_list:
        fail_file_path = os.path.join(download_path, "failed_downloads.txt")
        with open(fail_file_path, "w", encoding="utf-8") as f:
            for url in fail_url_list:
                f.write(url + "\n")
        print(f"❌ {len(fail_url_list)} 个种子下载失败，已保存到 {fail_file_path}")


def 批量获取数据(url_list: list[str]) -> Tuple[List[dict], List[dict]]:
    """
    批量获取动画信息和单集信息
    :param url_list: 包含多个 Bangumi URL 的列表
    :return: 返回动画信息列表和单集信息列表
    """

    anime_info_list = []
    episode_info_list = []

    # 多线程实现
    with ThreadPoolExecutor() as executor:
        future_to_url = {executor.submit(utils.request_html, url): url for url in url_list}

        for future in tqdm(as_completed(future_to_url), total=len(future_to_url), desc="获取番组数据进度"):
            url = future_to_url[future]
            try:
                html_str = future.result()
                anime_info, episode_info = bangumi.解析BangumiHTML_str(html_str)
                anime_info_list.append(anime_info)
                episode_info_list.extend(episode_info)
            except Exception as e:
                print(f"❌ 获取 {url} 时发生错误: {e}")

    # 返回动画信息和单集信息
    return anime_info_list, episode_info_list


def 批量获取种子数据(data: dict[str, str]) -> list[dict]:
    """
    批量获取种子数据
    :param data: 包含多个番组链接和对应 RSS 订阅链接的字典列表 即 番组链接 : RSS订阅链接
    :return: 返回所有种子数据的列表
    """

    # 使用多线程批量获取种子数据
    种子数据列表: list[dict] = []
    with ThreadPoolExecutor() as executor:

        def 获取种子数据(bgm_url: str, mikan_rss_url: str) -> list[dict]:

            if mikan_rss_url is None:
                return []
            try:
                rss_html_str = utils.request_html(mikan_rss_url)
            except Exception as e:
                print(f"❌ 获取 {bgm_url}: {mikan_rss_url} 时发生错误: {e}")
                return []

            return mikananime.解析mikanRSS_XML(bgm_url, rss_html_str)

        futures = {
            executor.submit(获取种子数据, bgm_url, rss_url): (bgm_url, rss_url) for bgm_url, rss_url in data.items()
        }

        for future in tqdm(as_completed(futures), total=len(futures), desc="获取种子数据进度"):
            try:
                result = future.result()
                if result:
                    种子数据列表.extend(result)
            except Exception as e:
                print(f"❌ 获取种子数据时发生错误: {e}")

    # 返回所有种子数据
    return 种子数据列表


# 同步数据到 Access 数据库
def 更新数据库(data: list[dict], pk: str, headers_no_pk: list[str], accdb_path: str, table_name: str):
    """
    同步数据到 Access 数据库
    :param data:     list[dict]，每一行为一个字典，可能包含无关字段
    :param headers:  需要写入的字段列表（顺序指定）
    :param accdb_path: Access 数据库路径
    :param table_name: 目标表名

    逻辑：
    - 将 headers 的第一列作为主键
    - 遍历数据：
        - 如果缺主键或主键值为空，跳过
        - 若主键已存在 → 仅更新 headers 中指定的字段
        - 否则 → 仅插入 headers 中指定的字段
    """

    def database_print(msg: str, end: str = "\n"):
        print(f"\033[92m[数据库操作]:\033[0m {msg}", end=end)

    database_print(f"同步数据到数据库: {accdb_path} 的表 {table_name} : ", "")

    conn_str = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};" rf"DBQ={accdb_path};"
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    # 1. 获取主键列名
    if not pk:
        raise ValueError("❌ 主键列名 pk 不能为空")
    elif not headers_no_pk or len(headers_no_pk) == 0:
        raise ValueError("❌ headers 列表不能为空")
    pk_column = pk

    if not pk_column:
        raise Exception(f"❌ 无法获取 Access 表 [{table_name}] 的主键列")

    插入_count = 0
    更新_count = 0

    for record in data:
        if pk_column not in record or not record[pk_column]:
            database_print(f"⚠️ 跳过记录，缺少主键 [{pk_column}]：{record}")
            continue

        pk_value = record[pk_column]

        # 2. 判断主键是否存在
        cursor.execute(f"SELECT COUNT(*) FROM [{table_name}] WHERE [{pk_column}] = ?", (pk_value,))
        exists = cursor.fetchone()[0] > 0

        if exists:
            # 3. 执行更新
            update_fields = ", ".join(f"[{h}] = ?" for h in headers_no_pk)
            update_sql = f"UPDATE [{table_name}] SET {update_fields} WHERE [{pk_column}] = ?"
            update_values = [record.get(h, "") for h in headers_no_pk]
            cursor.execute(update_sql, tuple(update_values) + (pk_value,))
            更新_count += 1

        else:
            # 4. 执行插入
            field_names = ", ".join(f"[{h}]" for h in headers_no_pk)
            field_names += f", [{pk_column}]"  # 添加主键列
            placeholders = ", ".join("?" for _ in headers_no_pk)
            placeholders += ", ?"
            insert_sql = f"INSERT INTO [{table_name}] ({field_names}) VALUES ({placeholders})"
            insert_values = [record.get(h, "") for h in headers_no_pk]
            cursor.execute(insert_sql, tuple(insert_values) + (pk_value,))
            插入_count += 1

    conn.commit()
    cursor.close()
    conn.close()

    print("同步完成")
    database_print(f"➕ 插入记录数：{插入_count}")
    database_print(f"🔄 更新记录数：{更新_count}")
    print()


def 读取EXCEL并更新数据库(EXCEL文件地址):

    kumigumiPrint(f"📖 读取 Excel 文件: {EXCEL文件地址}")

    wb = load_workbook(EXCEL文件地址, data_only=True)
    sheet_main = wb["main"]

    数据库地址: str = ""
    数据库anime表名: str = ""
    数据库episode表名: str = ""
    数据库torrent表名: str = ""

    excel_anime_sheet_fetch_list: list[str] = []
    excel_anime_sheet_store_list: list[str] = []
    excel_episode_sheet_store_list: list[str] = []
    excel_torrent_sheet_store_list: list[str] = []

    要下载的种子的状态: str = ""
    torrent_download_sheet_name: str = ""  # 用于存储种子下载链接的工作表名
    torrent_download_url_list: list[str] = []  # 用于存储种子下载链接

    # 解析 main 工作表
    行指针: int = 1
    while True:
        cell_Ax = sheet_main.cell(行指针, 1).value

        if cell_Ax == "_end":
            break
        elif cell_Ax == "_to":  # 跳到指定行
            行指针 = sheet_main.cell(行指针, 2).value
            continue
        elif cell_Ax is None:
            pass

        elif cell_Ax == "_database_path":
            数据库地址 = sheet_main.cell(行指针, 2).value
        elif cell_Ax == "_anime_table":
            数据库anime表名 = sheet_main.cell(行指针, 2).value
        elif cell_Ax == "_episode_table":
            数据库episode表名 = sheet_main.cell(行指针, 2).value
        elif cell_Ax == "_torrent_table":
            数据库torrent表名 = sheet_main.cell(行指针, 2).value

        elif cell_Ax == "_download_torrent":
            torrent_download_sheet_name = sheet_main.cell(行指针, 2).value
            要下载的种子的状态 = sheet_main.cell(行指针, 3).value

        elif cell_Ax == "_store":
            数据库表类型 = sheet_main.cell(行指针, 2).value
            工作表名 = sheet_main.cell(行指针, 3).value
            if 数据库表类型 == "_anime_table":
                excel_anime_sheet_store_list.append(工作表名)
            elif 数据库表类型 == "_episode_table":
                excel_episode_sheet_store_list.append(工作表名)
            elif 数据库表类型 == "_torrent_table":
                excel_torrent_sheet_store_list.append(工作表名)

        elif cell_Ax == "_fetch":
            excel_anime_sheet_fetch_list.append(sheet_main.cell(行指针, 3).value)

        else:
            kumigumiPrint(f"⚠️ 未知指令: {cell_Ax}")

        行指针 += 1

    # 检查是否定义变量
    if not 数据库地址 or not 数据库anime表名 or not 数据库episode表名 or not 数据库torrent表名:
        raise ValueError("❌ 请确保在 main 工作表中定义了数据库地址和表名")

    # 更新 Access 数据库
    for 数据库表名, 工作表名_list in zip(
        [数据库anime表名, 数据库episode表名, 数据库torrent表名],
        [excel_anime_sheet_store_list, excel_episode_sheet_store_list, excel_torrent_sheet_store_list],
    ):
        for 工作表名 in 工作表名_list:
            kumigumiPrint("🔄 更新 Access 数据库...")
            sheet_download_torrent = wb[工作表名]

            起始行: int = 0
            结束行: int = 0
            主键: str = ""
            字段字典: dict[str, int] = {}  # 字段名 : 列号

            行指针: int = 1
            while True:
                键: str = sheet_download_torrent.cell(row=行指针, column=1).value
                值: str = sheet_download_torrent.cell(row=行指针, column=2).value

                if 键 is None:
                    pass
                elif cell_Ax == "_to":  # 跳到指定行
                    行指针 = int(值)
                    continue
                elif 键 == "_end":
                    break
                elif 键 == "_start_row":
                    起始行 = int(值)
                elif 键 == "_end_row":
                    结束行 = int(值)
                elif 键 == "_primary_key":
                    主键 = 值
                else:
                    字段字典[键] = int(值)

                行指针 += 1

            # 翻译
            主键 = headers.字段字典.get(主键, 主键)
            字段字典 = {headers.字段字典.get(k, k): v for k, v in 字段字典.items()}

            # 读取数据区域
            data: list[dict[str, int]] = []
            for 行号 in range(起始行, 结束行):
                row_data: dict[str, int] = {}
                for 字段名, 列号 in 字段字典.items():
                    单元格值 = sheet_download_torrent.cell(row=行号, column=列号).value
                    row_data[字段名] = 单元格值 if 单元格值 is not None else ""
                data.append(row_data)

            # 更新 Access 数据库
            更新数据库(data, 主键, [k for k in 字段字典.keys() if k != 主键], 数据库地址, 数据库表名)

    # 批量获取远程数据并更新数据库
    for 源sheet in excel_anime_sheet_fetch_list:
        kumigumiPrint("🔄 批量获取远程数据并更新数据库...")

        bgm_url_column: int = 0
        rss_url_column: int = 0
        起始行: int = 0
        结束行: int = 0

        # 读取源工作表
        print(f"📖 读取源工作表: {源sheet}")
        sheet_download_torrent = wb[源sheet]
        行指针 = 1
        while True:
            cell_Ax = sheet_download_torrent.cell(行指针, 1).value

            # 仅获取番组链接和RSS订阅链接
            if cell_Ax == "_end":
                break
            elif cell_Ax == "_to":  # 跳到指定行
                行指针 = int(sheet_download_torrent.cell(行指针, 2).value)
                continue
            elif cell_Ax is None:
                pass
            elif cell_Ax == "_start_row":
                起始行 = int(sheet_download_torrent.cell(行指针, 2).value)
            elif cell_Ax == "_end_row":
                结束行 = int(sheet_download_torrent.cell(行指针, 2).value)
            elif cell_Ax == "番组bangumi链接":
                bgm_url_column = sheet_download_torrent.cell(行指针, 2).value
            elif cell_Ax == "番组RSS订阅链接":
                rss_url_column = sheet_download_torrent.cell(行指针, 2).value

            行指针 += 1

        # 读取信息
        bgm_url_rss_映射: dict[str, str] = {}  # 番组链接 : RSS订阅链接
        for 行号 in range(起始行, 结束行):
            bgm_url = sheet_download_torrent.cell(行号, bgm_url_column).value
            rss_url = sheet_download_torrent.cell(行号, rss_url_column).value
            bgm_url_rss_映射[bgm_url] = rss_url

        anime_info_list, episode_info_list = 批量获取数据(bgm_url_rss_映射.keys())
        torrent_info_list = 批量获取种子数据(bgm_url_rss_映射)

        # 翻译键名
        anime_info_list = [{headers.字段字典.get(k, k): v for k, v in row.items()} for row in anime_info_list]
        episode_info_list = [{headers.字段字典.get(k, k): v for k, v in row.items()} for row in episode_info_list]
        torrent_info_list = [{headers.字段字典.get(k, k): v for k, v in row.items()} for row in torrent_info_list]

        kumigumiPrint("获取完毕")

        # 同步动画信息到 Access
        更新数据库(
            anime_info_list,
            headers.番组表头_主键_en,
            headers.番组表头_自动更新_en,
            数据库地址,
            数据库anime表名,
        )
        更新数据库(
            episode_info_list,
            headers.单集表头_主键_en,
            headers.单集表头_自动更新_en,
            数据库地址,
            数据库episode表名,
        )
        更新数据库(
            torrent_info_list,
            headers.种子表头_主键_en,
            headers.种子表头_自动更新_en,
            数据库地址,
            数据库torrent表名,
        )

    # 下载种子链接
    if torrent_download_sheet_name != "":
        kumigumiPrint("🔄 下载种子链接...")

        # 获取种子下载链接工作表
        sheet_download_torrent = wb[torrent_download_sheet_name]

        起始行: int = 0
        结束行: int = 0
        种子下载链接_column: int = 0
        种子下载情况_column: int = 0

        # 读取种子下载链接
        行指针 = 1
        while True:
            cell_Ax = sheet_download_torrent.cell(行指针, 1).value

            if cell_Ax == "_end":
                break
            elif cell_Ax == "_to":
                行指针 = int(sheet_download_torrent.cell(行指针, 2).value)
                continue
            elif cell_Ax is None:
                pass

            elif cell_Ax == "_start_row":
                起始行 = int(sheet_download_torrent.cell(行指针, 2).value)
            elif cell_Ax == "_end_row":
                结束行 = int(sheet_download_torrent.cell(行指针, 2).value)

            elif cell_Ax == "种子下载链接":
                种子下载链接_column = sheet_download_torrent.cell(行指针, 2).value
            elif cell_Ax == "种子下载情况":
                种子下载情况_column = sheet_download_torrent.cell(行指针, 2).value

            行指针 += 1

        # 读取种子下载链接
        for 行号 in range(起始行, 结束行):
            torrent_download_url = sheet_download_torrent.cell(行号, 种子下载链接_column).value
            torrent_download_status = sheet_download_torrent.cell(行号, 种子下载情况_column).value
            if torrent_download_status == 要下载的种子的状态:
                torrent_download_url_list.append(torrent_download_url)

        批量下载种子(torrent_download_url_list)


def safe_load_excel(path) -> str:
    """
    创建一个临时文件，复制指定的 Excel 文件到临时文件中，
    然后使用 openpyxl 加载临时文件以避免文件被占用
    """

    temp_path = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(path, temp_path)

    return temp_path


if __name__ == "__main__":

    warnings.filterwarnings("ignore", category=UserWarning)

    kumigumiPrint("开始执行脚本...")

    excel_path = "D:/OneDrive/kumigumi.xlsx"
    读取EXCEL并更新数据库(safe_load_excel(excel_path))

    kumigumiPrint("所有操作完成")
