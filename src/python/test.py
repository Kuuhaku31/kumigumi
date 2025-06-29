# test.py


from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Tuple

import pyodbc
from openpyxl import load_workbook
from tqdm import tqdm

import bangumi
import headers
import mikananime.mikananime as mikananime
import utils


def 批量获取数据(url_list: list[str]) -> Tuple[List[dict], List[dict]]:
    """
    批量获取动画信息和单集信息
    :param url_list: 包含多个 Bangumi URL 的列表
    :return: 返回动画信息列表和单集信息列表
    """

    anime_info_list = []
    episode_info_list = []

    # 多线程实现
    with ThreadPoolExecutor() as executor:
        future_to_url = {executor.submit(utils.request_html, url): url for url in url_list}

        for future in tqdm(as_completed(future_to_url), total=len(future_to_url), desc="获取进度"):
            url = future_to_url[future]
            try:
                html_str = future.result()
                anime_info, episode_info = bangumi.解析BangumiHTML_str(html_str)
                anime_info_list.append(anime_info)
                episode_info_list.extend(episode_info)
            except Exception as e:
                print(f"❌ 获取 {url} 时发生错误: {e}")

    # 返回动画信息和单集信息
    return anime_info_list, episode_info_list


def 批量获取种子数据(data: list[dict]) -> List[dict]:

    # 使用多线程批量获取种子数据
    种子数据列表: list[dict] = []
    with ThreadPoolExecutor() as executor:

        def 获取种子数据(bgm_url: str, mikan_rss_url: str) -> list[dict]:
            try:
                rss_html_str = utils.request_html(mikan_rss_url)
            except Exception as e:
                print(f"❌ 获取 {bgm_url}: {mikan_rss_url} 时发生错误: {e}")
                return []

            return mikananime.解析mikanRSS_XML(bgm_url, rss_html_str)

        futures = {
            executor.submit(获取种子数据, row["番组bangumi链接"], row["番组RSS订阅链接"]): row
            for row in data
            if row.get("番组bangumi链接") and row.get("番组RSS订阅链接")
        }

        for future in tqdm(as_completed(futures), total=len(futures), desc="获取种子数据进度"):
            try:
                result = future.result()
                种子数据列表.extend(result)
            except Exception as e:
                print(f"❌ 获取种子数据时发生错误: {e}")

    # 返回所有种子数据
    return 种子数据列表


def 创建数据库表(accdb_path: str, table_name: str, headers: list[str], overwrite: bool = False):
    """
    创建 Access 数据库表
    :param accdb_path: Access 数据库路径
    :param table_name: 要创建的表名
    :param headers: 表头列表（第一个字段作为主键）
    :param overwrite: 是否覆盖已存在的表
    """
    if not headers:
        raise ValueError("❌ 表头列表不能为空")

    conn_str = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};" rf"DBQ={accdb_path};"
    conn = pyodbc.connect(conn_str, autocommit=True)
    cursor = conn.cursor()

    # 删除旧表（如果 overwrite = True）
    if overwrite:
        try:
            cursor.execute(f"DROP TABLE [{table_name}]")
            print(f"🗑️ 已删除旧表：{table_name}")
        except Exception as e:
            print(f"⚠️ 删除旧表失败或不存在：{e}")

    # 构建字段定义
    field_defs = [f"[{headers[0]}] TEXT PRIMARY KEY"]  # 主键字段
    field_defs += [f"[{col}] TEXT" for col in headers[1:]]

    create_sql = f"CREATE TABLE [{table_name}] ({', '.join(field_defs)})"
    cursor.execute(create_sql)
    print(f"✅ 已创建新表：{table_name}，字段数：{len(headers)}")

    cursor.close()
    conn.close()


# 同步数据到 Access 数据库
def 更新数据库(data: list[dict], pk: str, headers_no_pk: list[str], accdb_path: str, table_name: str):
    """
    同步数据到 Access 数据库
    :param data:     list[dict]，每一行为一个字典，可能包含无关字段
    :param headers:  需要写入的字段列表（顺序指定）
    :param accdb_path: Access 数据库路径
    :param table_name: 目标表名

    逻辑：
    - 将 headers 的第一列作为主键
    - 遍历数据：
        - 如果缺主键或主键值为空，跳过
        - 若主键已存在 → 仅更新 headers 中指定的字段
        - 否则 → 仅插入 headers 中指定的字段
    """

    print(f"🔄 同步数据到 Access 数据库: {accdb_path} 的表 {table_name}")

    conn_str = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};" rf"DBQ={accdb_path};"
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    # 1. 获取主键列名
    #  将 headers 的第一列作为主键

    if not pk:
        raise ValueError("❌ 主键列名 pk 不能为空")
    elif not headers_no_pk or len(headers_no_pk) == 0:
        raise ValueError("❌ headers 列表不能为空")

    pk_column = pk

    # cursor.execute(f"SELECT * FROM [{table_name}]")
    # pk_column = None
    # for column in cursor.description:
    #     if column[5]:  # column[5] 为 True 表示是主键
    #         pk_column = column[0]
    #         break

    if not pk_column:
        raise Exception(f"❌ 无法获取 Access 表 [{table_name}] 的主键列")

    插入_count = 0
    更新_count = 0

    for record in data:
        if pk_column not in record or not record[pk_column]:
            print(f"⚠️ 跳过记录，缺少主键 [{pk_column}]：{record}")
            continue

        # 仅保留 headers 中字段，按顺序提取值（空填""）
        # row = [record.get(h, "") for h in headers]
        pk_value = record[pk_column]

        # 2. 判断主键是否存在
        cursor.execute(f"SELECT COUNT(*) FROM [{table_name}] WHERE [{pk_column}] = ?", (pk_value,))
        exists = cursor.fetchone()[0] > 0

        if exists:
            # 3. 执行更新
            update_fields = ", ".join(f"[{h}] = ?" for h in headers_no_pk)
            update_sql = f"UPDATE [{table_name}] SET {update_fields} WHERE [{pk_column}] = ?"
            update_values = [record.get(h, "") for h in headers_no_pk]
            cursor.execute(update_sql, tuple(update_values) + (pk_value,))
            更新_count += 1

        else:
            # 4. 执行插入
            field_names = ", ".join(f"[{h}]" for h in headers_no_pk)
            field_names += f", [{pk_column}]"  # 添加主键列
            placeholders = ", ".join("?" for _ in headers_no_pk)
            placeholders += ", ?"
            insert_sql = f"INSERT INTO [{table_name}] ({field_names}) VALUES ({placeholders})"
            insert_values = [record.get(h, "") for h in headers_no_pk]
            cursor.execute(insert_sql, tuple(insert_values) + (pk_value,))
            插入_count += 1

    conn.commit()
    cursor.close()
    conn.close()

    print("✅ 同步完成")
    print(f"➕ 插入记录数：{插入_count}")
    print(f"🔄 更新记录数：{更新_count}")


def 读取EXCEL表格区域(excel_file_path: str, sheet_name: str) -> list[dict]:
    """
    读取 Excel 表格中指定工作表的区域（由 A1:B* 定义）
    返回：数据字典列表

    Excel 表格中的 A1:B* 区域格式为：
    |      | A               | B
    | 1    | len             | 需要返回数据的字段的个数（即 A4:B* 的行数）
    | 2    | start_row       | 数据区域开始的行号（不包括表头，即第一行数据所在的行的行号）
    | 3    | end_row         | 最后一条数据的下一行的行号
    | 4    | 种子下载链接    | 该字段所在的列号
    | 5    | 种子下载情况    | 该字段所在的列号
    | 6    | 备注            | 该字段所在的列号
    | ...  | ...             | ...

    """

    print(f"📖 读取 Excel 文件: {excel_file_path} 的工作表: {sheet_name}")

    wb = load_workbook(excel_file_path, data_only=True)
    sheet = wb[sheet_name]

    # 读取元信息部分
    index_len = sheet.cell(row=1, column=2).value
    start_row = sheet.cell(row=2, column=2).value
    end_row = sheet.cell(row=3, column=2).value

    字段信息 = {}
    row = 4
    while row < 4 + index_len:
        key_cell = sheet.cell(row=row, column=1).value
        value_cell = sheet.cell(row=row, column=2).value
        字段信息[key_cell] = value_cell
        row += 1

    # 读取数据区域
    result = []
    for 行号 in range(start_row, end_row):
        row_data = {}
        for field, col in 字段信息.items():
            cell_value = sheet.cell(row=行号, column=col).value
            row_data[field] = cell_value
            if row_data[field] is None:
                row_data[field] = ""

        result.append(row_data)

    return result


# Access 数据库路径和表名
全局_accdb_path = "D:/def/test_db.accdb"
全局_数据库anime表名 = "anime"
全局_数据库episode表名 = "episode"
全局_数据库torrent表名 = "torrent"
# 全局_kumigumi_db_path = "D:/def/kumigumi.accdb"

excel_path = "D:/def/2025.07.xlsx"
excel_sheet_name = "ani_index"
excel_sheet_name_torrent_db = "torrent_db"
excel_sheet_name_ep202504 = "ep202504"
excel_sheet_name_ani202507 = "ani202507"
excel_sheet_name_ani202504 = "ani202504"


def 读取表格区域并爬取数据然后更新数据库(EXCEL文件地址, 工作表名):
    print("读取表格区域并爬取数据然后更新数据库")

    # 读取 Excel 表格区域
    data = 读取EXCEL表格区域(EXCEL文件地址, 工作表名)

    bgm_url_list: list[str] = []
    for row in data:
        if row.get("番组bangumi链接"):
            bgm_url_list.append(row["番组bangumi链接"])

    # 批量获取数据
    anime_info, episode_info = 批量获取数据(bgm_url_list)

    # 翻译键名
    anime_info = [{headers.字段字典.get(k, k): v for k, v in row.items()} for row in anime_info]
    episode_info = [{headers.字段字典.get(k, k): v for k, v in row.items()} for row in episode_info]

    # 同步动画信息到 Access
    更新数据库(
        anime_info,
        headers.番组表头_主键_en,
        headers.番组表头_自动更新_en,
        全局_accdb_path,
        全局_数据库anime表名,
    )
    更新数据库(
        episode_info,
        headers.单集表头_主键_en,
        headers.单集表头_自动更新_en,
        全局_accdb_path,
        全局_数据库episode表名,
    )


def 读取表格数据并爬取种子信息然后保存到数据库(EXCEL文件地址, 工作表名):
    print("读取表格数据并爬取种子信息然后保存到数据库")

    # 读取 Excel 表格区域
    data = 读取EXCEL表格区域(EXCEL文件地址, 工作表名)

    # 批量获取种子数据
    data = 批量获取种子数据(data)
    data = [{headers.字段字典.get(k, k): v for k, v in row.items()} for row in data]

    # 同步种子数据到 Access
    更新数据库(
        data,
        headers.种子表头_主键_en,
        headers.种子表头_自动更新_en,
        全局_accdb_path,
        全局_数据库torrent表名,
    )


def 读取EXCEL并更新数据库(EXCEL文件地址):

    print(f"📖 读取 Excel 文件: {EXCEL文件地址}")

    wb = load_workbook(EXCEL文件地址, data_only=True)
    sheet_main = wb["main"]

    数据库地址: str = ""
    数据库表名_sheet_映射: dict[str, str] = {}  # 数据库表名 : 工作表名

    行指针: int = 1
    while True:
        cell_Ax = sheet_main.cell(行指针, 1).value

        if cell_Ax == "_end":
            break
        elif cell_Ax is None:
            pass
        elif cell_Ax == "_database_path":
            数据库地址 = sheet_main.cell(行指针, 2).value
        elif cell_Ax == "_store":
            数据库表名 = sheet_main.cell(行指针, 2).value
            工作表名 = sheet_main.cell(行指针, 3).value
            数据库表名_sheet_映射[数据库表名] = 工作表名
        else:
            print(f"⚠️ 未知指令: {cell_Ax}")

        行指针 += 1

    # 更新 Access 数据库
    for 数据库表名, 工作表名 in 数据库表名_sheet_映射.items():
        print(f"🔄 更新数据库: {数据库地址} 的表 {数据库表名}，工作表名: {工作表名}")

        sheet = wb[工作表名]
        起始行: int = 0
        结束行: int = 0
        主键: str = ""
        字段字典: dict[str, int] = {}  # 字段名 : 列号

        键: str = ""
        值: str = ""
        行指针: int = 1
        while True:
            键 = sheet.cell(row=行指针, column=1).value
            值 = sheet.cell(row=行指针, column=2).value

            if 键 is None:
                pass
            elif 键 == "_end":
                break
            elif 键 == "_start_row":
                起始行 = int(值)
            elif 键 == "_end_row":
                结束行 = int(值)
            elif 键 == "_primary_key":
                主键 = 值
            else:
                字段字典[键] = int(值)

            行指针 += 1

        # 翻译
        主键 = headers.字段字典.get(主键, 主键)
        字段字典 = {headers.字段字典.get(k, k): v for k, v in 字段字典.items()}

        print(f"数据库地址: {数据库地址}")
        print(f"数据库表名: {数据库表名}")
        print(f"起始行: {起始行}")
        print(f"结束行: {结束行}")
        print(f"主键: {主键}")
        print(f"字段字典: {字段字典}")

        # 读取数据区域
        data = []
        for 行号 in range(起始行, 结束行):
            row_data = {}
            for 字段名, 列号 in 字段字典.items():
                单元格值 = sheet.cell(row=行号, column=列号).value
                row_data[字段名] = 单元格值 if 单元格值 is not None else ""
            data.append(row_data)

        # 更新 Access 数据库
        headers_no_pk = [k for k in 字段字典.keys() if k != 主键]
        更新数据库(data, 主键, headers_no_pk, 数据库地址, 数据库表名)


if __name__ == "__main__":

    print("开始执行脚本...")

    读取EXCEL并更新数据库(excel_path)
    # 读取表格区域并更新数据库(excel_path, excel_sheet_name_ani202504, "a")
    # 读取表格区域并爬取数据然后更新数据库(excel_path, excel_sheet_name_ani202507)
    # 读取表格数据并爬取种子信息然后保存到数据库(excel_path, "ani202504")

    print("所有操作完成")
