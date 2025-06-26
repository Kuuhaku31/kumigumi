# to_excel.py

import os
import sqlite3

import pandas as pd
import pyodbc
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def sqlite_table_to_excel(
    sqlite_path: str,
    table_name: str,
    excel_path: str,
    sheet_name: str = "Sheet1",
    start_row: int = 0,
    start_col: int = 0,
):
    # 1. 从 SQLite 读取表为 DataFrame
    conn = sqlite3.connect(sqlite_path)
    try:
        df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
    except Exception as e:
        print(f"读取 SQLite 表失败：{e}")
        conn.close()
        return
    conn.close()

    # 2. 打开或创建 Excel 文件
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        from openpyxl import Workbook

        wb = Workbook()

    # 3. 获取或创建指定 Sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # 4. 写入数据（含列名）从指定位置开始
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row):
            ws.cell(row=start_row + r_idx + 1, column=start_col + c_idx + 1, value=value)

    # 5. 保存 Excel 文件
    wb.save(excel_path)
    print(f"成功将表 [{table_name}] 写入 {excel_path} → Sheet:{sheet_name}, 从({start_row+1}, {start_col+1}) 开始")


# CSV → ACCDB
def csv_to_accdb_table(csv_path, accdb_path, table_name, overwrite=False):
    # 读取 CSV（防止空值变成 NaN，我们后面手动处理）
    df = pd.read_csv(csv_path, dtype=str).fillna("")  # 将 NaN 替换为 ""

    # 连接 Access 数据库
    conn_str = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};" rf"DBQ={accdb_path};"
    conn = pyodbc.connect(conn_str, autocommit=True)
    cursor = conn.cursor()

    # 表存在时是否覆盖
    if overwrite:
        try:
            cursor.execute(f"DROP TABLE [{table_name}]")
            print(f"✅ 已删除旧表：{table_name}")
        except Exception as e:
            print(f"⚠️ 删除旧表失败：{table_name}，可能不存在或有其他问题，错误信息: {e}")
            pass

    # 创建表结构（全部字段 TEXT 类型）
    col_defs = ", ".join(f"[{col}] TEXT" for col in df.columns)
    create_sql = f"CREATE TABLE [{table_name}] ({col_defs})"
    cursor.execute(create_sql)
    print(f"✅ 创建新表：{table_name}")

    # 插入数据（空值已在读取时填成 ""）
    for _, row in df.iterrows():
        placeholders = ", ".join(["?"] * len(row))
        insert_sql = f"INSERT INTO [{table_name}] VALUES ({placeholders})"
        cursor.execute(insert_sql, tuple(row))  # 无需转换类型，已是 str + ""

    conn.commit()
    cursor.close()
    conn.close()

    print(f"✅ 成功将 {csv_path} 导入为 {accdb_path} 中的表 [{table_name}]")


def find_true_rows_by_named_range(excel_path, named_range, target_column_name) -> tuple[list[str], list[list]]:
    wb = load_workbook(excel_path, data_only=True, keep_links=False)

    # 获取命名区域
    if named_range not in wb.defined_names:
        print(f"❌ 命名区域 '{named_range}' 不存在")
        return [], []

    # 获取该命名区域对应的表格位置
    defined_range = wb.defined_names[named_range]

    print("命名区域内容：", defined_range.attr_text)

    dest = list(defined_range.destinations)[0]  # 只取第一个（一般只定义一个区域）
    sheet_name, cell_range = dest
    ws = wb[sheet_name]

    # 获取起始行和列范围
    from openpyxl.utils import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    # 获取表头
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col, max_col + 1)]

    if target_column_name not in headers:
        print(f"❌ 表头中找不到列名 '{target_column_name}'")
        return [], []

    # 获取目标列的索引
    target_col_index = headers.index(target_column_name)

    matched_rows = []
    for row in range(min_row + 1, max_row + 1):  # 从数据行开始（跳过表头）
        row_values = [ws.cell(row=row, column=col).value for col in range(min_col, max_col + 1)]

        # 检查目标列的值是否为 True
        val = row_values[target_col_index]
        if isinstance(val, bool) and val is True:
            row_values[target_col_index - 4] = row_values[target_col_index - 2]
            row_values[target_col_index - 3] = row_values[target_col_index - 1]
            matched_rows.append(row_values[:-3])

    return headers[:-3], matched_rows


# rows: 二维列表（每一行是一个记录）
# headers: 表头（字段名列表）
# accdb_path: Access 数据库路径
# table_name: 要更新的目标表名
def sync_rows_to_access(rows, headers, accdb_path, table_name):
    conn_str = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};" rf"DBQ={accdb_path};"
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    primary_key = headers[0]  # 默认第一列是主键

    for row in rows:
        pk_value = row[0]

        # 1. 检查是否已存在该主键记录
        cursor.execute(f"SELECT COUNT(*) FROM [{table_name}] WHERE [{primary_key}] = ?", (pk_value,))
        exists = cursor.fetchone()[0] > 0

        # 2. 执行更新
        if exists:
            update_fields = ", ".join(f"[{h}] = ?" for h in headers[1:])
            update_sql = f"UPDATE [{table_name}] SET {update_fields} WHERE [{primary_key}] = ?"
            cursor.execute(update_sql, tuple(row[1:]) + (pk_value,))
            print(f"🔄 更新：{pk_value}")

        # 3. 执行插入
        else:
            field_names = ", ".join(f"[{h}]" for h in headers)
            placeholders = ", ".join("?" for _ in headers)
            insert_sql = f"INSERT INTO [{table_name}] ({field_names}) VALUES ({placeholders})"
            cursor.execute(insert_sql, tuple(row))
            print(f"➕ 插入：{pk_value}")

    conn.commit()
    cursor.close()
    conn.close()
    print("✅ 同步完成")


if __name__ == "__main__":

    # 示例用法
    sqlite_path = "./test.db"
    table_name = "torrents"  # 替换为你的 SQLite 表名
    excel_path = "test.xlsx"
    sheet_name = "TorrentsData"

    accdb_path = "./test_db.accdb"

    csv_path = "./episode.csv"
    csv_path_2 = "D:/OneDrive/kumigumi/2025.07/episode.csv"

    table_name_accdb = "EpisodesData"

    target_column_name = "need_update"
    named_range = "ep_data"

    # 将 SQLite 表导出到 Excel
    # sqlite_table_to_excel(sqlite_path, table_name, excel_path, sheet_name)

    # 将 CSV 导入到 Access 数据库
    # csv_to_accdb_table(csv_path_2, accdb_path, table_name_accdb, overwrite=True)

    # exit(0)

    # 查找 Excel 中命名区域内指定列为 True 的行
    headers, matched_rows = find_true_rows_by_named_range(excel_path, named_range, target_column_name)

    if matched_rows and headers:
        # 打印表头
        print("表头：", headers)
        # 打印找到的行
        print(f"在命名区域 '{named_range}' 中找到 {len(matched_rows)} 行 '{target_column_name}' 为 True 的数据：")
        for row in matched_rows:
            print(row)

        # 同步到 Access 数据库
        sync_rows_to_access(matched_rows, headers, accdb_path, table_name_accdb)

    else:
        print(f"在命名区域 '{named_range}' 中没有找到 '{target_column_name}' 为 True 的数据。")
